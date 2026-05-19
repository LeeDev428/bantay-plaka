"""
Shared export helpers — Excel (openpyxl) and PDF (reportlab).
Usage:
    from apps.core.export_utils import ExcelExporter, PdfExporter
"""
import datetime
from django.http import HttpResponse
from django.utils import timezone


# ── Shared style constants ────────────────────────────────────────────────────

NAVY   = '1E3A5F'
BLUE   = '2563EB'
ALT    = 'F1F5F9'
WHITE  = 'FFFFFF'
RED_BG = 'FEE2E2'
SLATE  = '334155'
LIGHT  = 'BFDBFE'


class ExcelExporter:
    """
    Build a styled .xlsx response.

    Usage:
        ex = ExcelExporter('my_export', date_from, date_to)
        ws = ex.add_sheet('Data', ['Col A', 'Col B'], col_widths=[20, 30])
        ex.write_rows(ws, rows, highlight_fn=lambda r: r[1] == 'BAD')
        ex.add_summary_sheet('Summary', [('Total', 100), ('Flagged', 3)])
        return ex.response()
    """

    def __init__(self, base_name: str, date_from='', date_to='', total_records=0):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError('openpyxl is not installed. Run: pip install openpyxl')

        self._openpyxl = openpyxl
        self._Font = Font
        self._PatternFill = PatternFill
        self._Alignment = Alignment
        self._Border = Border
        self._Side = Side
        self._col_letter = get_column_letter

        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)   # remove default sheet; we'll add explicitly
        self.base_name = base_name
        self.date_from = date_from
        self.date_to = date_to
        self.total_records = total_records

        # Build shared styles once
        thin = Side(style='thin', color='CBD5E1')
        self._thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self._center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self._left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        self._header_fill  = PatternFill('solid', fgColor=NAVY)
        self._subhead_fill = PatternFill('solid', fgColor=BLUE)
        self._alt_fill     = PatternFill('solid', fgColor=ALT)
        self._white_fill   = PatternFill('solid', fgColor=WHITE)
        self._red_fill     = PatternFill('solid', fgColor=RED_BG)

    # ── helpers ───────────────────────────────────────────────────────────────

    def _apply(self, cell, font=None, fill=None, align=None, border=None):
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border

    def _title_font(self, size=14, color=WHITE, bold=True):
        return self._Font(name='Calibri', bold=bold, size=size, color=color)

    def _body_font(self, bold=False, color=SLATE, size=9):
        return self._Font(name='Calibri', bold=bold, size=size, color=color)

    def _mono_font(self, color=NAVY):
        return self._Font(name='Courier New', size=9, bold=True, color=color)

    # ── public API ────────────────────────────────────────────────────────────

    def add_sheet(self, title: str, headers: list, col_widths: list = None) -> object:
        """Add a sheet with title block + header row. Returns the worksheet."""
        ws = self.wb.create_sheet(title)
        ncols = len(headers)
        last_col = self._col_letter(ncols)

        # Row 1 — title
        ws.merge_cells(f'A1:{last_col}1')
        c = ws['A1']
        c.value = title.upper()
        self._apply(c,
            font=self._title_font(size=14),
            fill=self._header_fill,
            align=self._center,
            border=self._thin_border,
        )
        ws.row_dimensions[1].height = 28

        # Row 2 — meta
        ws.merge_cells(f'A2:{last_col}2')
        date_label = ''
        if self.date_from and self.date_to:
            date_label = f'  |  Period: {self.date_from} to {self.date_to}'
        elif self.date_from:
            date_label = f'  |  From: {self.date_from}'
        elif self.date_to:
            date_label = f'  |  To: {self.date_to}'
        c = ws['A2']
        c.value = f'Generated: {timezone.localdate()}{date_label}  |  Total: {self.total_records}'
        c.font = self._Font(name='Calibri', size=9, color=LIGHT)
        c.fill = self._header_fill
        c.alignment = self._center
        ws.row_dimensions[2].height = 14

        # Row 3 — spacer
        ws.row_dimensions[3].height = 5

        # Row 4 — column headers
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=i, value=h)
            self._apply(c,
                font=self._Font(name='Calibri', bold=True, size=10, color=WHITE),
                fill=self._subhead_fill,
                align=self._center,
                border=self._thin_border,
            )
        ws.row_dimensions[4].height = 20

        # Column widths
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[self._col_letter(i)].width = w

        # Freeze panes
        ws.freeze_panes = 'A5'
        return ws

    def write_rows(self, ws, rows: list,
                   highlight_fn=None,
                   mono_cols: set = None,
                   center_cols: set = None):
        """
        Write data rows starting at row 5.
        highlight_fn(row_data) -> True means red background.
        mono_cols: 1-based column indices to render in monospace.
        center_cols: 1-based column indices to center-align.
        """
        mono_cols   = mono_cols   or set()
        center_cols = center_cols or set()
        ncols = ws.max_column or len(rows[0]) if rows else 1

        for r_idx, row in enumerate(rows, 1):
            excel_row = r_idx + 4
            flagged = highlight_fn(row) if highlight_fn else False
            fill = self._red_fill if flagged else (self._alt_fill if r_idx % 2 == 0 else self._white_fill)

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=excel_row, column=c_idx, value=value)
                cell.fill = fill
                cell.border = self._thin_border
                cell.alignment = self._center if c_idx in center_cols else self._left

                if c_idx == 1:
                    cell.font = self._Font(name='Calibri', size=8, color='94A3B8')
                elif c_idx in mono_cols:
                    cell.font = self._Font(name='Courier New', size=9, bold=True,
                                           color='DC2626' if flagged else NAVY)
                else:
                    cell.font = self._body_font(color='DC2626' if flagged else SLATE)

            ws.row_dimensions[excel_row].height = 15

        # Auto-filter
        if rows:
            ws.auto_filter.ref = f'A4:{self._col_letter(len(rows[0]))}{4 + len(rows)}'

    def add_summary_sheet(self, title: str, items: list):
        """Add a summary sheet with label/value pairs."""
        ws = self.wb.create_sheet(title)
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 16

        ws.merge_cells('A1:B1')
        c = ws['A1']
        c.value = 'SUMMARY'
        self._apply(c,
            font=self._title_font(size=12),
            fill=self._header_fill,
            align=self._center,
        )
        ws.row_dimensions[1].height = 22

        for i, (label, value) in enumerate(items, 2):
            fill = self._alt_fill if i % 2 == 0 else self._white_fill
            lc = ws.cell(row=i, column=1, value=label)
            vc = ws.cell(row=i, column=2, value=value)
            lc.fill = vc.fill = fill
            lc.font = self._body_font(size=10, color=SLATE)
            vc.font = self._body_font(size=10, bold=True, color=NAVY)
            lc.alignment = self._left
            vc.alignment = self._center
            lc.border = vc.border = self._thin_border
            ws.row_dimensions[i].height = 18

    def response(self) -> HttpResponse:
        filename = f'{self.base_name}_{timezone.localdate()}.xlsx'
        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        self.wb.save(resp)
        return resp


class PdfExporter:
    """
    Build a styled landscape PDF response via reportlab.
    Usage:
        pdf = PdfExporter('my_export', title='My Report', date_from=..., date_to=...)
        pdf.build(headers, rows, col_widths)
        return pdf.response()
    """

    def __init__(self, base_name: str, title: str, date_from='', date_to=''):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            raise ImportError('reportlab is not installed. Run: pip install reportlab')

        self._colors = colors
        self._mm = mm
        self._Table = Table
        self._TableStyle = TableStyle
        self._Paragraph = Paragraph
        self._Spacer = Spacer
        self._styles = getSampleStyleSheet()
        self._ParagraphStyle = ParagraphStyle
        self._TA_CENTER = TA_CENTER
        self._TA_LEFT = TA_LEFT

        self.base_name = base_name
        self.title = title
        self.date_from = date_from
        self.date_to = date_to
        self._buffer = HttpResponse(content_type='application/pdf')
        filename = f'{base_name}_{timezone.localdate()}.pdf'
        self._buffer['Content-Disposition'] = f'attachment; filename="{filename}"'

        self.doc = SimpleDocTemplate(
            self._buffer,
            pagesize=landscape(A4),
            leftMargin=12*mm, rightMargin=12*mm,
            topMargin=12*mm, bottomMargin=12*mm,
        )
        self.elements = []
        self._build_header()

    def _build_header(self):
        mm = self._mm
        colors = self._colors
        P = self._Paragraph
        S = self._Spacer
        styles = self._styles

        title_style = self._ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.HexColor(f'#{NAVY}'),
            spaceAfter=2,
        )
        sub_style = self._ParagraphStyle(
            'CustomSub',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#64748B'),
        )
        self.elements.append(P(self.title, title_style))

        date_label = ''
        if self.date_from and self.date_to:
            date_label = f' &nbsp;|&nbsp; Period: {self.date_from} to {self.date_to}'
        elif self.date_from:
            date_label = f' &nbsp;|&nbsp; From: {self.date_from}'
        elif self.date_to:
            date_label = f' &nbsp;|&nbsp; To: {self.date_to}'
        self.elements.append(P(f'Generated: {timezone.localdate()}{date_label}', sub_style))
        self.elements.append(S(1, 5*mm))

    def build(self, headers: list, rows: list, col_widths: list,
              highlight_fn=None, mono_col_idx: int = None):
        """
        headers: list of str
        rows: list of lists
        col_widths: list of mm values
        highlight_fn(row) -> True for red highlight
        mono_col_idx: 0-based column to bold (e.g. plate number)
        """
        colors = self._colors
        mm = self._mm

        data = [headers]
        for row in rows:
            data.append([str(v) if v is not None else '' for v in row])

        col_widths_pts = [w * mm for w in col_widths]
        table = self._Table(data, colWidths=col_widths_pts, repeatRows=1)

        style_cmds = [
            ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor(f'#{NAVY}')),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 8),
            ('FONTSIZE',     (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor(f'#{ALT}')]),
            ('GRID',         (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',   (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
            ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
        ]

        if highlight_fn:
            for i, row in enumerate(rows, 1):
                if highlight_fn(row):
                    style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor(f'#{RED_BG}')))
                    style_cmds.append(('TEXTCOLOR',  (0, i), (-1, i), colors.HexColor('#DC2626')))

        if mono_col_idx is not None:
            style_cmds.append(('FONTNAME', (mono_col_idx, 1), (mono_col_idx, -1), 'Courier-Bold'))

        table.setStyle(self._TableStyle(style_cmds))
        self.elements.append(table)

    def response(self) -> HttpResponse:
        self.doc.build(self.elements)
        return self._buffer