"""
Reusable export helpers for Excel (.xlsx) and PDF.
Drop this file anywhere and import directly — no Django app dependency.

    from path.to.export_helpers import build_excel_response, build_pdf_response
"""
from django.http import HttpResponse
from django.utils import timezone

NAVY  = '1E3A5F'
BLUE  = '2563EB'
ALT   = 'F1F5F9'
WHITE = 'FFFFFF'
REDBG = 'FEE2E2'
SLATE = '334155'
LIGHT = 'BFDBFE'


def build_excel_response(filename_base, sheet_title, headers, col_widths,
                         rows, highlight_fn=None, mono_cols=None,
                         center_cols=None, summary_rows=None,
                         date_from='', date_to=''):
    """
    Returns an HttpResponse with an .xlsx attachment.
    rows: list of lists (first column should be row #)
    highlight_fn(row) -> bool  — red background when True
    mono_cols / center_cols: 1-based sets of column indices
    summary_rows: list of (label, value) for a second Summary sheet
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mono_cols   = mono_cols   or set()
    center_cols = center_cols or set()

    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def fill(hex_): return PatternFill('solid', fgColor=hex_)
    def font(bold=False, size=9, color=SLATE, name='Calibri'):
        return Font(name=name, bold=bold, size=size, color=color)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ncols = len(headers)
    last = get_column_letter(ncols)

    # ── Title ──
    ws.merge_cells(f'A1:{last}1')
    c = ws['A1']
    c.value = sheet_title.upper()
    c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    c.fill = fill(NAVY); c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 28

    # ── Meta ──
    ws.merge_cells(f'A2:{last}2')
    c = ws['A2']
    date_label = ''
    if date_from and date_to:   date_label = f'  |  Period: {date_from} – {date_to}'
    elif date_from:             date_label = f'  |  From: {date_from}'
    elif date_to:               date_label = f'  |  To: {date_to}'
    c.value = f'Generated: {timezone.localdate()}  |  Total: {len(rows)}{date_label}'
    c.font = Font(name='Calibri', size=9, color=LIGHT)
    c.fill = fill(NAVY); c.alignment = center
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 5  # spacer

    # ── Column headers ──
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
        c.fill = fill(BLUE); c.alignment = center; c.border = border
    ws.row_dimensions[4].height = 20

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ──
    for r_idx, row in enumerate(rows, 1):
        erow = r_idx + 4
        flagged = highlight_fn(row) if highlight_fn else False
        bg = fill(REDBG) if flagged else (fill(ALT) if r_idx % 2 == 0 else fill(WHITE))
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=erow, column=c_idx, value=val)
            cell.fill = bg; cell.border = border
            cell.alignment = center if c_idx in center_cols else left
            if c_idx == 1:
                cell.font = Font(name='Calibri', size=8, color='94A3B8')
            elif c_idx in mono_cols:
                cell.font = Font(name='Courier New', size=9, bold=True,
                                 color='DC2626' if flagged else NAVY)
            else:
                cell.font = Font(name='Calibri', size=9,
                                 color='DC2626' if flagged else SLATE)
        ws.row_dimensions[erow].height = 15

    if rows:
        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f'A4:{last}{4 + len(rows)}'

    # ── Summary sheet ──
    if summary_rows:
        ws2 = wb.create_sheet('Summary')
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 16
        ws2.merge_cells('A1:B1')
        c = ws2['A1']
        c.value = 'SUMMARY'
        c.font = Font(name='Calibri', bold=True, size=12, color=WHITE)
        c.fill = fill(NAVY); c.alignment = center
        ws2.row_dimensions[1].height = 22
        for i, (label, value) in enumerate(summary_rows, 2):
            bg = fill(ALT) if i % 2 == 0 else fill(WHITE)
            lc = ws2.cell(row=i, column=1, value=label)
            vc = ws2.cell(row=i, column=2, value=value)
            lc.fill = vc.fill = bg
            lc.font = Font(name='Calibri', size=10, color=SLATE)
            vc.font = Font(name='Calibri', size=10, bold=True, color=NAVY)
            lc.alignment = left; vc.alignment = center
            lc.border = vc.border = border
            ws2.row_dimensions[i].height = 18

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename_base}_{timezone.localdate()}.xlsx"'
    wb.save(resp)
    return resp


def build_pdf_response(filename_base, report_title, headers, col_widths_mm,
                       rows, highlight_fn=None, mono_col_idx=None,
                       date_from='', date_to=''):
    """
    Returns an HttpResponse with a landscape A4 PDF attachment.
    col_widths_mm: list of mm values matching headers length.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="{filename_base}_{timezone.localdate()}.pdf"'
    )
    doc = SimpleDocTemplate(resp, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=15,
                                 textColor=colors.HexColor(f'#{NAVY}'), spaceAfter=2)
    sub_style   = ParagraphStyle('S', parent=styles['Normal'], fontSize=8,
                                 textColor=colors.HexColor('#64748B'))

    date_label = ''
    if date_from and date_to:   date_label = f' | Period: {date_from} – {date_to}'
    elif date_from:             date_label = f' | From: {date_from}'
    elif date_to:               date_label = f' | To: {date_to}'

    elems = [
        Paragraph(report_title, title_style),
        Paragraph(f'Generated: {timezone.localdate()}  |  Total: {len(rows)}{date_label}', sub_style),
        Spacer(1, 5*mm),
    ]

    data = [headers] + [[str(v) if v is not None else '' for v in r] for r in rows]
    table = Table(data, colWidths=[w*mm for w in col_widths_mm], repeatRows=1)

    style_cmds = [
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor(f'#{NAVY}')),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 8),
        ('FONTSIZE',      (0,1), (-1,-1), 7),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor(f'#{ALT}')]),
        ('GRID',          (0,0), (-1,-1), 0.3, colors.HexColor('#CBD5E1')),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('ALIGN',         (0,0), (-1,0),  'CENTER'),
    ]
    if highlight_fn:
        for i, row in enumerate(rows, 1):
            if highlight_fn(row):
                style_cmds += [
                    ('BACKGROUND', (0,i), (-1,i), colors.HexColor(f'#{REDBG}')),
                    ('TEXTCOLOR',  (0,i), (-1,i), colors.HexColor('#DC2626')),
                ]
    if mono_col_idx is not None:
        style_cmds.append(('FONTNAME', (mono_col_idx,1), (mono_col_idx,-1), 'Courier-Bold'))

    table.setStyle(TableStyle(style_cmds))
    elems.append(table)
    doc.build(elems)
    return resp