import csv
import datetime
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Subquery, OuterRef
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from django.shortcuts import render
from django.utils import timezone

from apps.logs.models import VehicleLog
from apps.logs.services import get_active_blacklist_map


def _day_range(date):
    """Return (start_utc, end_utc) for a local date, avoiding MySQL CONVERT_TZ dependency."""
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.datetime.combine(date, datetime.time.min), tz)
    end = start + timedelta(days=1)
    return start, end


@login_required
def report_dashboard(request):
    if request.user.is_resident():
        messages.error(request, 'Access denied.')
        return redirect('resident_dashboard')

    today = timezone.localdate()
    today_start, today_end = _day_range(today)
    today_logs = VehicleLog.objects.filter(timestamp__gte=today_start, timestamp__lt=today_end)
    today_in = today_logs.filter(status=VehicleLog.STATUS_IN).count()
    today_out = today_logs.filter(status=VehicleLog.STATUS_OUT).count()
    today_unique = today_logs.values('plate_number').distinct().count()

    # Visitors currently inside: latest log is TIME_IN and latest type is VISITOR.
    latest_status = (
        VehicleLog.objects
        .filter(plate_number=OuterRef('plate_number'))
        .order_by('-timestamp')
        .values('status')[:1]
    )
    latest_type = (
        VehicleLog.objects
        .filter(plate_number=OuterRef('plate_number'))
        .order_by('-timestamp')
        .values('entry_type')[:1]
    )
    currently_inside = (
        VehicleLog.objects
        .values('plate_number')
        .distinct()
        .annotate(last_status=Subquery(latest_status))
        .annotate(last_type=Subquery(latest_type))
        .filter(last_status=VehicleLog.STATUS_IN, last_type=VehicleLog.TYPE_VISITOR)
        .count()
    )
    currently_inside_items = list(
        VehicleLog.objects
        .values('plate_number')
        .distinct()
        .annotate(last_status=Subquery(latest_status))
        .annotate(last_type=Subquery(latest_type))
        .filter(last_status=VehicleLog.STATUS_IN, last_type=VehicleLog.TYPE_VISITOR)
        .order_by('plate_number')
    )

    # 7-day daily breakdown — use UTC ranges so no CONVERT_TZ needed
    daily_data = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        d_start, d_end = _day_range(d)
        day_qs = VehicleLog.objects.filter(timestamp__gte=d_start, timestamp__lt=d_end)
        daily_data.append({
            'date': d.strftime('%Y-%m-%d'),
            'label': d.strftime('%a'),
            'time_in': day_qs.filter(status=VehicleLog.STATUS_IN).count(),
            'time_out': day_qs.filter(status=VehicleLog.STATUS_OUT).count(),
        })

    # Top vehicles with filters and pagination
    top_window_q = request.GET.get('top_window', 'WEEK').upper()
    top_entry_type_q = request.GET.get('top_entry_type', 'ALL').upper()
    if top_window_q == 'MONTH':
        period_start = today - timedelta(days=30)
    else:
        top_window_q = 'WEEK'
        period_start = today - timedelta(days=today.weekday())
    period_start_dt, _ = _day_range(period_start)

    top_vehicles_qs = (
        VehicleLog.objects
        .filter(timestamp__gte=period_start_dt)
        .values('plate_number', 'entry_type')
        .annotate(visits=Count('id'))
        .order_by('-visits', 'plate_number')
    )
    if top_entry_type_q in {VehicleLog.TYPE_RESIDENT, VehicleLog.TYPE_VISITOR}:
        top_vehicles_qs = top_vehicles_qs.filter(entry_type=top_entry_type_q)
    else:
        top_entry_type_q = 'ALL'

    top_paginator = Paginator(top_vehicles_qs, 8)
    top_page_number = request.GET.get('top_page', 1)
    top_vehicles = top_paginator.get_page(top_page_number)
    top_rank_offset = (top_vehicles.number - 1) * top_paginator.per_page

    context = {
        'today': today,
        'today_iso': today.isoformat(),
        'today_in': today_in,
        'today_out': today_out,
        'today_unique': today_unique,
        'currently_inside': currently_inside,
        'currently_inside_items': currently_inside_items,
        'daily_data': daily_data,
        'top_vehicles': top_vehicles,
        'top_rank_offset': top_rank_offset,
        'top_window_q': top_window_q,
        'top_entry_type_q': top_entry_type_q,
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
def export_csv(request):
    if request.user.is_resident():
        messages.error(request, 'Access denied.')
        return redirect('resident_dashboard')

    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')
    logs = VehicleLog.objects.select_related('logged_by').order_by('-timestamp')

    tz = timezone.get_current_timezone()
    if date_from:
        try:
            dt_from = datetime.date.fromisoformat(date_from)
            start, _ = _day_range(dt_from)
            logs = logs.filter(timestamp__gte=start)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.date.fromisoformat(date_to)
            _, end = _day_range(dt_to)
            logs = logs.filter(timestamp__lt=end)
        except ValueError:
            pass

    response = HttpResponse(content_type='text/csv')
    filename = f'bantayplaka_logs_{timezone.localdate()}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Plate Number', 'Camera Role', 'Entry Type', 'Status', 'Source',
        'Resident/Visitor Name', 'Blacklist Tag', 'Blacklist Remarks', 'Logged By', 'Timestamp (Asia/Manila)',
    ])

    bl_map = get_active_blacklist_map([log.plate_number for log in logs])

    for log in logs:
        local_ts = timezone.localtime(log.timestamp)
        bl_info = bl_map.get((log.plate_number or '').upper(), {})
        writer.writerow([
            log.plate_number,
            log.camera_role,
            log.entry_type,
            log.status,
            log.source,
            log.resident_name or log.visitor_name or '',
            bl_info.get('tag', ''),
            bl_info.get('remarks', ''),
            log.logged_by.get_full_name() if log.logged_by else 'System',
            local_ts.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    return response


@login_required
def export_pdf(request):
    if request.user.is_resident():
        messages.error(request, 'Access denied.')
        return redirect('resident_dashboard')

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return HttpResponse('reportlab is not installed. Run: pip install reportlab', status=500)

    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')
    logs = VehicleLog.objects.select_related('logged_by').order_by('-timestamp')

    if date_from:
        try:
            dt_from = datetime.date.fromisoformat(date_from)
            start, _ = _day_range(dt_from)
            logs = logs.filter(timestamp__gte=start)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.date.fromisoformat(date_to)
            _, end = _day_range(dt_to)
            logs = logs.filter(timestamp__lt=end)
        except ValueError:
            pass

    response = HttpResponse(content_type='application/pdf')
    filename = f'bantayplaka_logs_{timezone.localdate()}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph('Bantay Plaka — Vehicle Log Report', styles['Title']))
    elements.append(Paragraph(f'Generated: {timezone.localdate()}', styles['Normal']))
    elements.append(Spacer(1, 6*mm))

    headers = ['Plate', 'Type', 'Status', 'Source', 'Name', 'Timestamp']
    data = [headers]

    bl_map = get_active_blacklist_map([log.plate_number for log in logs])

    for log in logs:
        local_ts = timezone.localtime(log.timestamp)
        name = log.resident_name or log.visitor_name or ''
        data.append([
            log.plate_number or '',
            log.entry_type,
            log.status,
            log.source,
            name,
            local_ts.strftime('%Y-%m-%d %H:%M'),
        ])

    col_widths = [35*mm, 25*mm, 22*mm, 22*mm, 60*mm, 40*mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def export_excel(request):
    if request.user.is_resident():
        messages.error(request, 'Access denied.')
        return redirect('resident_dashboard')

    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')
    logs = VehicleLog.objects.select_related('logged_by').order_by('-timestamp')

    if date_from:
        try:
            dt_from = datetime.date.fromisoformat(date_from)
            start, _ = _day_range(dt_from)
            logs = logs.filter(timestamp__gte=start)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.date.fromisoformat(date_to)
            _, end = _day_range(dt_to)
            logs = logs.filter(timestamp__lt=end)
        except ValueError:
            pass

    bl_map = get_active_blacklist_map([log.plate_number for log in logs])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vehicle Logs'

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill   = PatternFill('solid', fgColor='1E3A5F')
    subhead_fill  = PatternFill('solid', fgColor='2563EB')
    alt_fill      = PatternFill('solid', fgColor='F1F5F9')
    white_fill    = PatternFill('solid', fgColor='FFFFFF')
    red_fill      = PatternFill('solid', fgColor='FEE2E2')
    title_font    = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    sub_font      = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
    header_font   = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    body_font     = Font(name='Calibri', size=9, color='1E293B')
    mono_font     = Font(name='Courier New', size=9, bold=True, color='1E3A5F')
    red_font      = Font(name='Calibri', size=9, color='DC2626', bold=True)
    center        = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left          = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_side     = Side(style='thin', color='CBD5E1')
    thick_side    = Side(style='medium', color='1E3A5F')
    thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_border  = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

    # ── Title block (rows 1–3) ────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'BANTAY PLAKA — VEHICLE LOG REPORT'
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:J2')
    meta_cell = ws['A2']
    date_label = ''
    if date_from and date_to:
        date_label = f'  |  Period: {date_from} to {date_to}'
    elif date_from:
        date_label = f'  |  From: {date_from}'
    elif date_to:
        date_label = f'  |  To: {date_to}'
    meta_cell.value = f'Generated: {timezone.localdate()}{date_label}  |  Total Records: {logs.count()}'
    meta_cell.font = Font(name='Calibri', size=9, color='BFDBFE')
    meta_cell.fill = header_fill
    meta_cell.alignment = center
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # spacer

    # ── Column headers (row 4) ────────────────────────────────────────────────
    columns = [
        ('#',              6),
        ('Plate Number',  16),
        ('Entry Type',    13),
        ('Status',        12),
        ('Source',        11),
        ('Camera Role',   14),
        ('Name',          28),
        ('Blacklist Tag', 14),
        ('Blacklist Remarks', 30),
        ('Timestamp',     20),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = subhead_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[4].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, log in enumerate(logs, start=1):
        excel_row = row_idx + 4
        local_ts = timezone.localtime(log.timestamp)
        bl_info = bl_map.get((log.plate_number or '').upper(), {})
        is_blacklisted = bool(bl_info)
        is_alt = row_idx % 2 == 0

        fill = red_fill if is_blacklisted else (alt_fill if is_alt else white_fill)

        row_data = [
            row_idx,
            log.plate_number or '',
            log.get_entry_type_display() if hasattr(log, 'get_entry_type_display') else log.entry_type,
            'Time In' if log.status == VehicleLog.STATUS_IN else 'Time Out',
            log.get_source_display() if hasattr(log, 'get_source_display') else log.source,
            log.get_camera_role_display() if hasattr(log, 'get_camera_role_display') else log.camera_role,
            log.resident_name or log.visitor_name or '',
            bl_info.get('tag', ''),
            bl_info.get('remarks', ''),
            local_ts.strftime('%Y-%m-%d %H:%M:%S'),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center if col_idx in (1, 3, 4, 5, 6, 8, 10) else left

            if col_idx == 1:  # row number
                cell.font = Font(name='Calibri', size=8, color='94A3B8')
            elif col_idx == 2:  # plate
                cell.font = red_font if is_blacklisted else mono_font
            elif col_idx == 8 and is_blacklisted:  # blacklist tag
                cell.font = red_font
            else:
                cell.font = body_font

        ws.row_dimensions[excel_row].height = 15

    # ── Freeze panes & filter ─────────────────────────────────────────────────
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:J{4 + logs.count()}'

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 16

    summary_title = ws2.cell(row=1, column=1, value='SUMMARY')
    summary_title.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    summary_title.fill = header_fill
    summary_title.alignment = center
    ws2.merge_cells('A1:B1')
    ws2.row_dimensions[1].height = 24

    summary_rows = [
        ('Total Records', logs.count()),
        ('Time In', logs.filter(status=VehicleLog.STATUS_IN).count()),
        ('Time Out', logs.filter(status=VehicleLog.STATUS_OUT).count()),
        ('Residents', logs.filter(entry_type=VehicleLog.TYPE_RESIDENT).count()),
        ('Visitors', logs.filter(entry_type=VehicleLog.TYPE_VISITOR).count()),
        ('Camera Source', logs.filter(source=VehicleLog.SOURCE_CAMERA).count()),
        ('Manual Source', logs.filter(source=VehicleLog.SOURCE_MANUAL).count()),
        ('Blacklisted Plates', sum(1 for log in logs if (log.plate_number or '').upper() in bl_map)),
    ]
    for i, (label, value) in enumerate(summary_rows, start=2):
        lc = ws2.cell(row=i, column=1, value=label)
        vc = ws2.cell(row=i, column=2, value=value)
        row_fill = alt_fill if i % 2 == 0 else white_fill
        lc.fill = vc.fill = row_fill
        lc.font = Font(name='Calibri', size=10, color='334155')
        vc.font = Font(name='Calibri', size=10, bold=True, color='1E3A5F')
        lc.alignment = left
        vc.alignment = center
        lc.border = vc.border = thin_border
        ws2.row_dimensions[i].height = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'bantayplaka_logs_{timezone.localdate()}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_visitors_inside(request):
    if request.user.is_resident():
        messages.error(request, 'Access denied.')
        return redirect('resident_dashboard')

    latest_status = (
        VehicleLog.objects
        .filter(plate_number=OuterRef('plate_number'))
        .order_by('-timestamp')
        .values('status')[:1]
    )
    latest_type = (
        VehicleLog.objects
        .filter(plate_number=OuterRef('plate_number'))
        .order_by('-timestamp')
        .values('entry_type')[:1]
    )
    latest_ts = (
        VehicleLog.objects
        .filter(plate_number=OuterRef('plate_number'))
        .order_by('-timestamp')
        .values('timestamp')[:1]
    )
    latest_name = (
        VehicleLog.objects
        .filter(plate_number=OuterRef('plate_number'))
        .order_by('-timestamp')
        .values('visitor_name')[:1]
    )

    inside_plates = (
        VehicleLog.objects
        .values('plate_number')
        .distinct()
        .annotate(last_status=Subquery(latest_status))
        .annotate(last_type=Subquery(latest_type))
        .annotate(last_ts=Subquery(latest_ts))
        .annotate(last_name=Subquery(latest_name))
        .filter(last_status=VehicleLog.STATUS_IN, last_type=VehicleLog.TYPE_VISITOR)
        .order_by('plate_number')
    )

    tz = timezone.get_current_timezone()
    response = HttpResponse(content_type='text/csv')
    filename = f'bantayplaka_visitors_inside_{timezone.localdate()}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['Plate Number', 'Visitor Name', 'Entry Time (Asia/Manila)'])
    for item in inside_plates:
        local_ts = timezone.localtime(item['last_ts'], tz).strftime('%Y-%m-%d %H:%M:%S') if item['last_ts'] else ''
        writer.writerow([item['plate_number'], item['last_name'] or '', local_ts])

    return response